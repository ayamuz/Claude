"""Build structured Excel workbook from ACCME provider data.

Reads tilde-delimited provider records and creates a professional Excel
workbook with 5 sheets: All Providers, Tier 1 Targets, Mental Health Targets,
Spanish Market Focus, and High Volume.

Includes mental health enrichment: MH Relevance flag, Specialty Category,
Org Type, Global Footprint, and Pitch Angle columns.

Usage:
    python build_excel.py <input_data.txt> <output.xlsx>

Input format: One record per line, 16 tilde-delimited fields.
See SKILL.md for field definitions.
"""
import sys, html as htmlmod
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SPANISH_CITIES = ['miami','san antonio','los angeles','phoenix','el paso','tucson','albuquerque']
COLS = ['Provider Name','City','State','Country','Website','Accreditation Type',
        'Accreditation Status','Joint Providership','Activities/Year','Contact Name',
        'Contact Phone','Address','ZIP','Activity Formats','Accredited By','Provider ID',
        'Priority Tier','Spanish Market','High Volume','Commendation Status',
        'MH Relevance','Specialty Category','Org Type','Global Footprint','Pitch Angle','Notes']
COL_WIDTHS = [45,18,6,6,30,25,30,10,12,25,18,40,12,35,30,10,10,14,11,18,
              14,22,18,14,45,25]

ACC_TYPE_MAP = {'A': 'ACCME Accredited', 'J': 'Jointly Accredited', 'S': 'State Accredited'}
ACC_STATUS_MAP = {'C': 'Accreditation with Commendation', 'A': 'Accredited', 'P': 'Provisional',
                  'X': 'Probation', 'JA': 'Joint Accreditation', 'JC': 'Joint with Commendation', 'O': 'Other'}
JP_MAP = {'Y': 'Yes', 'N': 'No'}

# ── Mental health keyword categories ──
MH_CATEGORIES = {
    'Psychiatry': ['psychiatr', 'psych center', 'psych hospital'],
    'Mental Health': ['mental health', 'behavioral health', 'behavioral medicine'],
    'Neurology': ['neurol', 'neurosci', 'neuromod'],
    'Substance Use': ['substance', 'addiction', 'drug abuse', 'opioid', 'alcohol'],
    'Psychology': ['psycholog', 'psychother', 'counseling', 'counselling'],
    'Dementia/Cognitive': ['alzheimer', 'dementia', 'cognitive', 'memory disorder'],
    'Child/Adolescent': ['child', 'adolescent', 'pediatric hematology', 'pediatric mental',
                         'youth mental', "children's"],
    'Trauma/Crisis': ['trauma', 'ptsd', 'crisis', 'disaster'],
    'Sleep Medicine': ['sleep'],
    'Palliative/Hospice': ['hospice', 'palliative', 'end of life', 'end-of-life'],
    'Simulation': ['simulation in healthcare'],
    'Neurodiversity': ['autism', 'adhd', 'neurodiver', 'developmental disab'],
}
GLOBAL_KEYWORDS = ['international', 'global', 'world']
ORG_TYPE_RULES = [
    ('Professional Society', ['society', 'association', 'academy', 'college of', 'board of',
                               'institute of', 'council of']),
    ('Academic Medical Center', ['university', 'school of medicine', 'medical school',
                                  'college of medicine', 'medical center']),
    ('Government/Public Health', ['department of', 'va ', 'veterans', 'state medical',
                                   'public health', 'government']),
    ('Hospital/Health System', ['hospital', 'health system', 'health care system',
                                 'medical center', 'health network', 'healthcare']),
    ('Education/CME Company', ['education', 'cme', 'learning', 'continuing medical']),
]


def expand_codes(r):
    r['accreditation_type'] = ACC_TYPE_MAP.get(r['accreditation_type'], r['accreditation_type'])
    r['accreditation_status'] = ACC_STATUS_MAP.get(r['accreditation_status'], r['accreditation_status'])
    r['joint_providership'] = JP_MAP.get(r['joint_providership'], r['joint_providership'])


def classify_mh(name_lower, accredited_by_lower):
    """Return list of matching MH specialty categories."""
    combined = name_lower + ' ' + accredited_by_lower
    cats = []
    for cat, keywords in MH_CATEGORIES.items():
        for kw in keywords:
            if kw in combined:
                cats.append(cat)
                break
    return cats


def classify_org_type(name_lower):
    """Return the best-matching organization type."""
    for label, keywords in ORG_TYPE_RULES:
        for kw in keywords:
            if kw in name_lower:
                return label
    return 'Other'


def is_global(name_lower, country):
    """Check if provider has global/international scope."""
    if country and country not in ('USA', 'US', ''):
        return True
    for kw in GLOBAL_KEYWORDS:
        if kw in name_lower:
            return True
    return False


def generate_pitch(r):
    """Generate a suggested outreach angle based on provider profile."""
    parts = []
    cats = r.get('mh_categories', [])
    is_spanish = r.get('spanish') == 'Yes'
    is_global_flag = r.get('global_footprint') == 'Yes'
    activities = r.get('activities', 0)
    has_commendation = r.get('commendation') == 'Yes'

    if 'Psychiatry' in cats or 'Mental Health' in cats:
        parts.append('Psychiatric/MH CME content development')
    elif 'Neurology' in cats:
        parts.append('Neurology & brain health CME writing')
    elif 'Substance Use' in cats:
        parts.append('Addiction medicine educational content')
    elif 'Child/Adolescent' in cats:
        parts.append('Pediatric mental health curriculum')
    elif 'Dementia/Cognitive' in cats:
        parts.append('Cognitive health & dementia education')
    elif 'Trauma/Crisis' in cats:
        parts.append('Trauma-informed care CME')
    elif 'Neurodiversity' in cats:
        parts.append('Neurodiversity-focused education design')
    elif cats:
        parts.append(f'{cats[0]} CME content')

    if is_spanish:
        parts.append('Spanish-language adaptation')
    if is_global_flag:
        parts.append('Global mental health & cross-cultural adaptation')
    if activities >= 500 and not cats:
        parts.append('High-volume medical writing partnership')
    elif activities >= 100 and not cats:
        parts.append('Scalable CME content support')
    if has_commendation and not parts:
        parts.append('Premium CME content for commended program')

    return '; '.join(parts) if parts else ''


def compute_tier(r):
    activities = r.get('activities', 0)
    status = r.get('accreditation_status', '')
    state = r.get('state', '')
    city = (r.get('city', '') or '').lower()
    name = r.get('provider_name', '')
    acc_type = r.get('accreditation_type', '')
    has_commendation = 'Commendation' in status
    if activities >= 100 or has_commendation or state == 'PR' or city in SPANISH_CITIES:
        return 1
    if (activities >= 20 or acc_type == 'ACCME Accredited' or
        any(kw in name for kw in ['Medical Society','Medical Association','Academy','College'])):
        return 2
    return 3


def build(data_file, output_file):
    records = []
    with open(data_file) as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            parts = line.split('~')
            if len(parts) < 16:
                parts += [''] * (16 - len(parts))
            name = htmlmod.unescape(parts[0])
            activities = 0
            try:
                activities = int(parts[8])
            except:
                pass
            r = {
                'provider_name': name, 'city': parts[1], 'state': parts[2],
                'country': parts[3], 'website': parts[4], 'accreditation_type': parts[5],
                'accreditation_status': parts[6], 'joint_providership': parts[7],
                'activities': activities, 'contact_name': parts[9], 'contact_phone': parts[10],
                'address': parts[11], 'zip': parts[12], 'activity_formats': parts[13],
                'accredited_by': parts[14], 'provider_id': parts[15],
            }
            expand_codes(r)

            # Mental health enrichment
            name_lower = name.lower()
            accredited_by_lower = (r['accredited_by'] or '').lower()
            mh_cats = classify_mh(name_lower, accredited_by_lower)
            r['mh_categories'] = mh_cats
            r['mh_relevance'] = 'Yes' if mh_cats else 'No'
            r['specialty_category'] = ', '.join(mh_cats) if mh_cats else ''
            r['org_type'] = classify_org_type(name_lower)
            r['global_footprint'] = 'Yes' if is_global(name_lower, r['country']) else 'No'

            tier = compute_tier(r)
            city_lower = (r['city'] or '').lower()
            spanish = 'Yes' if (r['state'] == 'PR' or city_lower in SPANISH_CITIES) else 'No'
            high_vol = 'Yes' if activities >= 100 else 'No'
            commendation = 'Yes' if 'Commendation' in r['accreditation_status'] else 'No'
            r['tier'] = tier
            r['spanish'] = spanish
            r['high_vol'] = high_vol
            r['commendation'] = commendation

            # Generate pitch angle
            r['pitch_angle'] = generate_pitch(r)

            records.append(r)

    records.sort(key=lambda x: (x['tier'], -x['activities']))
    print(f"Total records: {len(records)}")
    t1 = [r for r in records if r['tier'] == 1]
    t2 = [r for r in records if r['tier'] == 2]
    t3 = [r for r in records if r['tier'] == 3]
    print(f"Tier 1: {len(t1)}, Tier 2: {len(t2)}, Tier 3: {len(t3)}")
    spanish_recs = [r for r in records if r['spanish'] == 'Yes']
    high_vol_recs = sorted([r for r in records if r['high_vol'] == 'Yes'], key=lambda x: -x['activities'])
    mh_recs = sorted([r for r in records if r['mh_relevance'] == 'Yes'],
                     key=lambda x: (x['tier'], -x['activities']))
    global_mh = [r for r in mh_recs if r['global_footprint'] == 'Yes' or r['spanish'] == 'Yes']
    print(f"Spanish Market: {len(spanish_recs)}, High Volume: {len(high_vol_recs)}")
    print(f"Mental Health Relevant: {len(mh_recs)}, Global/Spanish MH: {len(global_mh)}")

    # Count by MH specialty
    from collections import Counter
    cat_counts = Counter()
    for r in mh_recs:
        for c in r['mh_categories']:
            cat_counts[c] += 1
    print("\nMental Health by specialty:")
    for cat, cnt in cat_counts.most_common():
        print(f"  {cat}: {cnt}")

    wb = Workbook()
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='2F5496')
    mh_header_fill = PatternFill('solid', fgColor='7030A0')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_font = Font(name='Arial', size=10)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'))
    tier_fills = {1: PatternFill('solid', fgColor='C6EFCE'),
                  2: PatternFill('solid', fgColor='FFEB9C'),
                  3: PatternFill('solid', fgColor='FFC7CE')}
    mh_yes_fill = PatternFill('solid', fgColor='E2CFFC')

    def row_data(r):
        return [r['provider_name'], r['city'], r['state'], r['country'], r['website'],
                r['accreditation_type'], r['accreditation_status'], r['joint_providership'],
                r['activities'], r['contact_name'], r['contact_phone'], r['address'],
                r['zip'], r['activity_formats'], r['accredited_by'], r['provider_id'],
                r['tier'], r['spanish'], r['high_vol'], r['commendation'],
                r['mh_relevance'], r['specialty_category'], r['org_type'],
                r['global_footprint'], r['pitch_angle'], '']

    def write_sheet(ws, data, name, use_mh_header=False):
        ws.title = name
        ws.append(COLS)
        hfill = mh_header_fill if use_mh_header else header_fill
        for ci in range(1, len(COLS)+1):
            cell = ws.cell(row=1, column=ci)
            cell.font = header_font
            cell.fill = hfill
            cell.alignment = header_align
            cell.border = thin_border
        for ri, r in enumerate(data, start=2):
            rd = row_data(r)
            for ci, val in enumerate(rd, start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = data_font
                cell.border = thin_border
                if ci == 17:
                    cell.fill = tier_fills.get(val, PatternFill())
                    cell.alignment = Alignment(horizontal='center')
                if ci == 9:
                    cell.alignment = Alignment(horizontal='right')
                    cell.number_format = '#,##0'
                if ci == 21 and val == 'Yes':
                    cell.fill = mh_yes_fill
        for ci, w in enumerate(COL_WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = 'A2'

    write_sheet(wb.active, records, 'All Providers')
    ws2 = wb.create_sheet()
    write_sheet(ws2, t1, 'Tier 1 Targets')
    ws3 = wb.create_sheet()
    write_sheet(ws3, mh_recs, 'Mental Health Targets', use_mh_header=True)
    ws4 = wb.create_sheet()
    write_sheet(ws4, spanish_recs, 'Spanish Market Focus')
    ws5 = wb.create_sheet()
    write_sheet(ws5, high_vol_recs, 'High Volume')

    wb.save(output_file)
    print(f"\nSaved to {output_file}")

    # Stats
    states = {}
    for r in records:
        s = r['state'] or 'Unknown'
        states[s] = states.get(s, 0) + 1
    top_states = sorted(states.items(), key=lambda x: -x[1])[:10]
    print("\nTop 10 states:")
    for s, c in top_states:
        print(f"  {s}: {c}")

    acc_types = {}
    for r in records:
        t = r['accreditation_type'] or 'Unknown'
        acc_types[t] = acc_types.get(t, 0) + 1
    print("\nBy accreditation type:")
    for t, c in sorted(acc_types.items(), key=lambda x: -x[1]):
        print(f"  {t}: {c}")

    with_contact = sum(1 for r in records if r['contact_name'])
    with_website = sum(1 for r in records if r['website'])
    print(f"\nData completeness:")
    print(f"  Contact names: {with_contact}/{len(records)} ({100*with_contact//len(records)}%)")
    print(f"  Websites: {with_website}/{len(records)} ({100*with_website//len(records)}%)")

if __name__ == '__main__':
    data_file = sys.argv[1]
    output_file = sys.argv[2]
    build(data_file, output_file)
