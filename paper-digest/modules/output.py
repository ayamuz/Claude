"""
output.py — Excel spreadsheet generation for paper digest results.
"""

import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
    Border,
    Side,
)
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

logger = logging.getLogger(__name__)

HEADER_BG = "1F4E79"
HEADER_FG = "FFFFFF"
GRAY_ROW_BG = "D9D9D9"
GRAY_ROW_FG = "808080"
DOI_COLOR = "0563C1"

COLUMNS = [
    ("Rank",          8),
    ("Journal",       22),
    ("Title",         40),
    ("Authors",       30),
    ("DOI",           30),
    ("Published",     12),
    ("Abstract",      60),
    ("US Relevance",  13),
    ("Local Relevance", 15),
    ("Policy Score",  13),
    ("Accessibility", 13),
    ("Surprise Score", 14),
    ("Current Events", 14),
    ("Total Score",   12),
    ("Boosts",        30),
    ("Deprioritize",  30),
    ("Story Hook",    50),
]

SCORE_COLS = ["H", "I", "J", "K", "L", "M"]  # US Rel → Current Events
TOTAL_COL = "N"


def _make_header_fill() -> PatternFill:
    return PatternFill("solid", fgColor=HEADER_BG)


def _make_gray_fill() -> PatternFill:
    return PatternFill("solid", fgColor=GRAY_ROW_BG)


def _sort_papers(papers: list[dict]) -> list[dict]:
    """Sort papers: failed evaluations last, otherwise by total_score desc, then us_relevance desc."""
    def sort_key(p):
        if p.get("_eval_failed"):
            return (0, 0, 0)
        total = p.get("total_score") or 0
        us = p.get("us_relevance") or 0
        return (1, total, us)

    return sorted(papers, key=sort_key, reverse=True)


def write_spreadsheet(papers: list[dict], output_path: str, min_score: int) -> None:
    """
    Write a formatted .xlsx file with all evaluated papers.

    papers: list of paper dicts with evaluation fields populated
    output_path: file path for output
    min_score: papers with total_score < min_score get gray styling
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Paper Digest"

    # --- Header row ---
    header_font = Font(bold=True, color=HEADER_FG, name="Calibri")
    header_fill = _make_header_fill()
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # Enable auto-filter across all columns
    last_col = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col}1"

    # --- Sort papers ---
    sorted_papers = _sort_papers(papers)

    # --- Data rows ---
    gray_fill = _make_gray_fill()

    for rank, paper in enumerate(sorted_papers, start=1):
        row = rank + 1  # row 1 is header
        is_failed = paper.get("_eval_failed", False)
        total_score = paper.get("total_score")
        is_low_score = (
            not is_failed
            and total_score is not None
            and total_score < min_score
        )
        use_gray = is_low_score or is_failed

        # Build row values
        doi = paper.get("doi", "")
        boost_flags = paper.get("boost_flags") or []
        deprio_flags = paper.get("deprioritize_flags") or []

        row_values = [
            rank,
            paper.get("journal_name", ""),
            paper.get("title", ""),
            paper.get("authors", ""),
            doi,                                            # col E — hyperlink added below
            paper.get("published_date", ""),
            paper.get("abstract", "") or "",
            paper.get("us_relevance"),
            paper.get("local_relevance"),
            paper.get("policy_implications"),
            paper.get("accessibility"),
            paper.get("surprise_factor"),
            paper.get("current_events"),
            total_score,
            ", ".join(boost_flags),
            ", ".join(deprio_flags),
            paper.get("story_hook", ""),
        ]

        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)

            # Base alignment
            col_letter = get_column_letter(col_idx)
            if col_letter in SCORE_COLS or col_letter == TOTAL_COL:
                cell.alignment = Alignment(horizontal="center", vertical="top")
                if col_letter == TOTAL_COL and not is_failed:
                    cell.font = Font(bold=True, color=GRAY_ROW_FG if use_gray else "000000", name="Calibri")
            elif col_letter == "G":
                # Abstract — wrap text
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            elif col_letter == "E":
                # DOI — hyperlink
                if doi:
                    url = f"https://doi.org/{doi}"
                    cell.hyperlink = url
                    cell.value = doi
                    cell.font = Font(
                        color=GRAY_ROW_FG if use_gray else DOI_COLOR,
                        underline="single",
                        name="Calibri",
                    )
                cell.alignment = Alignment(vertical="top")
            else:
                cell.alignment = Alignment(vertical="top", wrap_text=False)

            # Gray styling for low-score or failed rows
            if use_gray:
                cell.fill = gray_fill
                # Don't overwrite DOI or Total Score font set above
                if col_letter not in ("E", TOTAL_COL):
                    current_font = cell.font
                    cell.font = Font(
                        color=GRAY_ROW_FG,
                        bold=current_font.bold,
                        underline=current_font.underline,
                        name="Calibri",
                    )

    # --- Conditional formatting: color scales on score columns ---
    score_color_scale = ColorScaleRule(
        start_type="num", start_value=1, start_color="FF0000",  # red
        mid_type="num",   mid_value=3,   mid_color="FFFF00",    # yellow
        end_type="num",   end_value=5,   end_color="00B050",    # green
    )
    total_color_scale = ColorScaleRule(
        start_type="num", start_value=6,  start_color="FF0000",
        mid_type="num",   mid_value=18,   mid_color="FFFF00",
        end_type="num",   end_value=30,   end_color="00B050",
    )

    last_data_row = len(sorted_papers) + 1

    for col_letter in SCORE_COLS:
        col_range = f"{col_letter}2:{col_letter}{last_data_row}"
        ws.conditional_formatting.add(col_range, score_color_scale)

    total_range = f"{TOTAL_COL}2:{TOTAL_COL}{last_data_row}"
    ws.conditional_formatting.add(total_range, total_color_scale)

    # --- Save ---
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info("Saved spreadsheet to %s", output_path)
